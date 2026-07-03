#!/usr/bin/env python3
"""
build_workbook.py — Prediction Market Operations Tracker (workbook stage)

Reads the CSV snapshot produced by fetch_data.py and builds
output/prediction_market_ops_tracker.xlsx with four sheets:

  * Markets   — one row per market (real snapshot)
  * Trades    — real trade activity for the top-10 sports markets, with
                LIVE Excel formula columns (XLOOKUP metadata, notional, whale flag)
  * Analysis  — LIVE SUMIFS/COUNTIFS/AVERAGEIFS/PERCENTILE formulas:
                thin-market risk, pricing sanity checks, activity trends
  * Dashboard — KPI formulas, conditional formatting, 2 native Excel charts

Every analytical value is a LIVE Excel formula (not a precomputed number), so the
workbook recalculates when opened or when the underlying data is edited. The build
is fully deterministic given a saved CSV snapshot; the snapshot date is written to
the Dashboard and documented in the README.

Usage:
    python3 build_workbook.py
    python3 build_workbook.py --markets data/raw/markets_latest.csv \
                              --trades  data/raw/trades_latest.csv
"""

from __future__ import annotations

import argparse
import json
import os
from datetime import datetime, date

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(ROOT, "data", "raw")
OUT_DIR = os.path.join(ROOT, "output")
OUT_PATH = os.path.join(OUT_DIR, "prediction_market_ops_tracker.xlsx")

# ---- styling constants ----------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SECTION_FONT = Font(bold=True, size=12, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="595959")
KPI_LABEL_FONT = Font(bold=True, size=11)
RED_FILL = PatternFill("solid", fgColor="FFC7CE")      # thin market
RED_FONT = Font(color="9C0006", bold=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")   # pricing discrepancy
YELLOW_FONT = Font(color="9C6500", bold=True)
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")   # whale
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_USD = "#,##0"
FMT_USD2 = "#,##0.00"
FMT_PRICE = "0.000"
FMT_DATE = "yyyy-mm-dd"
FMT_INT = "#,##0"


# --------------------------------------------------------------------------- #
# Snapshot loading
# --------------------------------------------------------------------------- #
def resolve_snapshot(args):
    """Return (markets_df, trades_df, snapshot_iso) honoring manifest / fallbacks."""
    manifest_path = os.path.join(RAW_DIR, "snapshot_manifest.json")
    snapshot_iso = None
    markets_path = args.markets
    trades_path = args.trades

    if not markets_path and os.path.exists(manifest_path):
        with open(manifest_path, encoding="utf-8") as f:
            man = json.load(f)
        snapshot_iso = man.get("snapshot_utc")
        markets_path = os.path.join(RAW_DIR, man.get("markets_file", "markets_latest.csv"))
        trades_path = os.path.join(RAW_DIR, man.get("trades_file", "trades_latest.csv"))

    # Fallbacks to committed sample snapshot.
    markets_path = markets_path or os.path.join(RAW_DIR, "markets_latest.csv")
    trades_path = trades_path or os.path.join(RAW_DIR, "trades_latest.csv")

    if not os.path.exists(markets_path):
        raise SystemExit(f"Markets CSV not found: {markets_path}\n"
                         f"Run: python3 fetch_data.py")

    markets = pd.read_csv(markets_path)
    trades = pd.read_csv(trades_path) if os.path.exists(trades_path) else pd.DataFrame()

    # Optional sportsbook odds (only present if fetch_data ran with a key set).
    odds_path = os.path.join(RAW_DIR, "odds_latest.csv")
    odds = pd.read_csv(odds_path) if os.path.exists(odds_path) else pd.DataFrame()

    if snapshot_iso is None:
        snapshot_iso = datetime.utcfromtimestamp(
            os.path.getmtime(markets_path)).replace(microsecond=0).isoformat()

    return markets, trades, odds, snapshot_iso


# --------------------------------------------------------------------------- #
# Small cell helpers
# --------------------------------------------------------------------------- #
def style_header_row(ws, row, first_col, last_col):
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def parse_date(v):
    if isinstance(v, str) and len(v) >= 10:
        try:
            return date.fromisoformat(v[:10])
        except ValueError:
            return None
    return None


# --------------------------------------------------------------------------- #
# Sheet: Markets
# --------------------------------------------------------------------------- #
def build_markets(ws, markets):
    headers = ["title", "category", "is_sports", "yes_price", "no_price",
               "volume_24h", "total_volume", "liquidity", "end_date",
               "days_to_resolution"]
    ws.append(headers)
    style_header_row(ws, 1, 1, len(headers))

    for _, m in markets.iterrows():
        ws.append([
            m["question"], m["category"], bool(m["is_sports"]),
            None if pd.isna(m["yes_price"]) else float(m["yes_price"]),
            None if pd.isna(m["no_price"]) else float(m["no_price"]),
            float(m["volume_24h"]), float(m["total_volume"]), float(m["liquidity"]),
            parse_date(m["end_date"]),
            None,  # days_to_resolution — filled with a live formula below
        ])

    n = len(markets)
    last = n + 1  # last data row (row 1 is header)
    for r in range(2, last + 1):
        # LIVE formula: days to resolution measured from the documented snapshot date.
        ws.cell(row=r, column=10).value = (
            f'=IF(I{r}="","",INT(I{r})-INT(SnapshotDate))')
        ws.cell(row=r, column=4).number_format = FMT_PRICE
        ws.cell(row=r, column=5).number_format = FMT_PRICE
        ws.cell(row=r, column=6).number_format = FMT_USD
        ws.cell(row=r, column=7).number_format = FMT_USD
        ws.cell(row=r, column=8).number_format = FMT_USD
        ws.cell(row=r, column=9).number_format = FMT_DATE
        ws.cell(row=r, column=10).number_format = FMT_INT

    widths = {"A": 52, "B": 12, "C": 10, "D": 10, "E": 10, "F": 13,
              "G": 15, "H": 14, "I": 12, "J": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    return last


# --------------------------------------------------------------------------- #
# Sheet: Trades
# --------------------------------------------------------------------------- #
def build_trades(ws, trades):
    headers = ["market", "date", "timestamp_utc", "side", "outcome", "price",
               "size_shares", "notional_usd", "mkt_category", "mkt_total_volume",
               "large_trade_flag"]
    ws.append(headers)
    style_header_row(ws, 1, 1, len(headers))

    for _, t in trades.iterrows():
        ts = None
        if isinstance(t.get("timestamp_iso"), str) and t["timestamp_iso"]:
            try:
                ts = datetime.fromisoformat(t["timestamp_iso"]).replace(tzinfo=None)
            except ValueError:
                ts = None
        ws.append([
            t["market"], parse_date(t.get("date")), ts,
            t.get("side", ""), t.get("outcome", ""),
            float(t["price"]), float(t["size_shares"]),
            None, None, None, None,  # F..K filled with live formulas below
        ])

    tn = len(trades)
    last = tn + 1
    for r in range(2, last + 1):
        # notional exposure = price * size (LIVE)
        ws.cell(row=r, column=8).value = f"=F{r}*G{r}"
        # market metadata pulled in via XLOOKUP from the Markets sheet (LIVE)
        ws.cell(row=r, column=9).value = f'=_xlfn.XLOOKUP(A{r},MktTitle,MktCat,"n/a")'
        ws.cell(row=r, column=10).value = f"=_xlfn.XLOOKUP(A{r},MktTitle,MktTotVol,0)"
        # whale flag: notional above the 95th-percentile threshold (LIVE)
        ws.cell(row=r, column=11).value = (
            f'=IF(H{r}>WhaleThreshold,"WHALE","")')
        ws.cell(row=r, column=2).number_format = FMT_DATE
        ws.cell(row=r, column=3).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=r, column=6).number_format = FMT_PRICE
        ws.cell(row=r, column=7).number_format = FMT_USD2
        ws.cell(row=r, column=8).number_format = FMT_USD2
        ws.cell(row=r, column=10).number_format = FMT_USD

    # Highlight whale trades.
    if tn:
        ws.conditional_formatting.add(
            f"K2:K{last}",
            CellIsRule(operator="equal", formula=['"WHALE"'],
                       fill=ORANGE_FILL, font=Font(color="C55A11", bold=True)))

    widths = {"A": 46, "B": 12, "C": 17, "D": 7, "E": 9, "F": 9, "G": 14,
              "H": 14, "I": 12, "J": 16, "K": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    return last


# --------------------------------------------------------------------------- #
# Sheet: Analysis
# --------------------------------------------------------------------------- #
def build_analysis(ws, markets, trades, m_last, t_last):
    n = len(markets)
    ws["A1"] = "ANALYSIS — live Excel formulas (recalculate on open / on data edit)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Every flag below is a real formula. Thresholds are editable in the "
                "block beneath — change one and the whole sheet re-evaluates.")
    ws["A2"].font = NOTE_FONT

    # ---- editable threshold block (rows 4-8) ------------------------------- #
    ws["A4"] = "High 24h-volume threshold (75th pct):"
    ws["B4"] = f"=PERCENTILE(Markets!$F$2:$F${m_last},0.75)"
    ws["A5"] = "Low liquidity threshold (25th pct):"
    ws["B5"] = f"=PERCENTILE(Markets!$H$2:$H${m_last},0.25)"
    ws["A6"] = "Book-sum deviation tolerance:"
    ws["B6"] = 0.02
    ws["A7"] = "Heavy-volume threshold (median 24h vol):"
    ws["B7"] = f"=MEDIAN(Markets!$F$2:$F${m_last})"
    ws["A8"] = "Whale threshold (95th pct notional):"
    ws["B8"] = (f"=PERCENTILE(Trades!$H$2:$H${t_last},0.95)"
                if t_last >= 2 else 0)
    for r in range(4, 9):
        ws.cell(row=r, column=1).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2).number_format = FMT_USD2 if r in (4, 5, 7, 8) else "0.00"
        ws.cell(row=r, column=2).font = Font(color="1F3864", bold=True)

    # ---- per-market risk & pricing table ----------------------------------- #
    hdr_row = 10
    headers = ["market", "category", "volume_24h", "liquidity", "thin_flag",
               "yes_price", "no_price", "book_sum", "abs_dev", "price_flag",
               "near_certain_active"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hdr_row, column=i, value=h)
    style_header_row(ws, hdr_row, 1, len(headers))
    ws["A9"] = "1 & 2 — LIQUIDITY RISK + PRICING SANITY (one row per tracked market)"
    ws["A9"].font = SECTION_FONT

    first = hdr_row + 1
    for i in range(n):
        r = first + i
        mr = 2 + i  # corresponding Markets data row
        ws.cell(row=r, column=1, value=f"=Markets!A{mr}")
        ws.cell(row=r, column=2, value=f"=Markets!B{mr}")
        ws.cell(row=r, column=3, value=f"=Markets!F{mr}").number_format = FMT_USD
        ws.cell(row=r, column=4, value=f"=Markets!H{mr}").number_format = FMT_USD
        # Thin market = high traffic AND shallow book (real platform-liability signal).
        ws.cell(row=r, column=5,
                value=f'=IF(AND(C{r}>=$B$4,D{r}<=$B$5),"THIN MARKET","")')
        ws.cell(row=r, column=6, value=f"=Markets!D{mr}").number_format = FMT_PRICE
        ws.cell(row=r, column=7, value=f"=Markets!E{mr}").number_format = FMT_PRICE
        ws.cell(row=r, column=8, value=f"=F{r}+G{r}").number_format = FMT_PRICE
        ws.cell(row=r, column=9, value=f"=ABS(H{r}-1)").number_format = FMT_PRICE
        # Book sum deviates from 1.00 beyond tolerance -> internal pricing discrepancy.
        ws.cell(row=r, column=10,
                value=f'=IF(I{r}>$B$6,"CHECK PRICING","")')
        # Near-certain (>0.95 or <0.05) yet still heavily traded -> worth review.
        ws.cell(row=r, column=11,
                value=f'=IF(AND(OR(F{r}>0.95,F{r}<0.05),C{r}>$B$7),'
                      f'"CERTAIN+ACTIVE","")')
    last_market_row = first + n - 1

    # Conditional formatting: red thin markets, yellow pricing discrepancies.
    ws.conditional_formatting.add(
        f"E{first}:E{last_market_row}",
        CellIsRule(operator="equal", formula=['"THIN MARKET"'],
                   fill=RED_FILL, font=RED_FONT))
    ws.conditional_formatting.add(
        f"J{first}:J{last_market_row}",
        CellIsRule(operator="equal", formula=['"CHECK PRICING"'],
                   fill=YELLOW_FILL, font=YELLOW_FONT))
    ws.conditional_formatting.add(
        f"K{first}:K{last_market_row}",
        CellIsRule(operator="equal", formula=['"CERTAIN+ACTIVE"'],
                   fill=YELLOW_FILL, font=YELLOW_FONT))

    # ---- 3 — activity trends (SUMIFS / COUNTIFS / AVERAGEIFS) --------------- #
    trend_row = last_market_row + 3
    ws.cell(row=trend_row, column=1,
            value="3 — ACTIVITY TRENDS (live SUMIFS / COUNTIFS / AVERAGEIFS)")
    ws.cell(row=trend_row, column=1).font = SECTION_FONT

    trange_h = f"Trades!$H$2:$H${t_last}"   # notional
    trange_b = f"Trades!$B$2:$B${t_last}"   # date
    trange_i = f"Trades!$I$2:$I${t_last}"   # mkt_category
    trange_d = f"Trades!$D$2:$D${t_last}"   # side

    # 3a — by day
    day_hdr = trend_row + 1
    day_headers = ["trade_date", "trade_volume_usd", "trade_count", "avg_trade_usd"]
    for i, h in enumerate(day_headers, start=1):
        ws.cell(row=day_hdr, column=i, value=h)
    style_header_row(ws, day_hdr, 1, 4)
    days = sorted({d for d in trades.get("date", pd.Series(dtype=str)).dropna().unique()}) \
        if len(trades) else []
    day_first = day_hdr + 1
    for i, dstr in enumerate(days):
        r = day_first + i
        ws.cell(row=r, column=1, value=parse_date(dstr)).number_format = FMT_DATE
        ws.cell(row=r, column=2,
                value=f"=SUMIFS({trange_h},{trange_b},A{r})").number_format = FMT_USD
        ws.cell(row=r, column=3, value=f"=COUNTIFS({trange_b},A{r})").number_format = FMT_INT
        ws.cell(row=r, column=4,
                value=f"=AVERAGEIFS({trange_h},{trange_b},A{r})").number_format = FMT_USD2
    day_last = day_first + max(len(days) - 1, 0)

    # 3b — by category
    cat_hdr = day_last + 2
    ws.cell(row=cat_hdr - 1, column=1, value="by market category").font = NOTE_FONT
    cat_headers = ["category", "trade_volume_usd", "trade_count", "avg_trade_usd"]
    for i, h in enumerate(cat_headers, start=1):
        ws.cell(row=cat_hdr, column=i, value=h)
    style_header_row(ws, cat_hdr, 1, 4)
    cats = sorted(markets.loc[markets["condition_id"].isin(
        trades.get("condition_id", pd.Series(dtype=str))), "category"].unique()) \
        if len(trades) else []
    if not cats and len(trades):
        cats = ["Soccer"]
    cat_first = cat_hdr + 1
    for i, cat in enumerate(cats):
        r = cat_first + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2,
                value=f'=SUMIFS({trange_h},{trange_i},A{r})').number_format = FMT_USD
        ws.cell(row=r, column=3, value=f"=COUNTIFS({trange_i},A{r})").number_format = FMT_INT
        ws.cell(row=r, column=4,
                value=f"=AVERAGEIFS({trange_h},{trange_i},A{r})").number_format = FMT_USD2
    cat_last = cat_first + max(len(cats) - 1, 0)

    # 3c — by side (BUY vs SELL) — order-flow imbalance signal
    side_hdr = cat_last + 2
    ws.cell(row=side_hdr - 1, column=1, value="by trade side (order-flow imbalance)").font = NOTE_FONT
    for i, h in enumerate(["side", "trade_volume_usd", "trade_count", "avg_trade_usd"], start=1):
        ws.cell(row=side_hdr, column=i, value=h)
    style_header_row(ws, side_hdr, 1, 4)
    side_first = side_hdr + 1
    for i, side in enumerate(["BUY", "SELL"]):
        r = side_first + i
        ws.cell(row=r, column=1, value=side)
        ws.cell(row=r, column=2,
                value=f'=SUMIFS({trange_h},{trange_d},A{r})').number_format = FMT_USD
        ws.cell(row=r, column=3, value=f"=COUNTIFS({trange_d},A{r})").number_format = FMT_INT
        ws.cell(row=r, column=4,
                value=f"=AVERAGEIFS({trange_h},{trange_d},A{r})").number_format = FMT_USD2

    # ---- manual pivot area ------------------------------------------------- #
    pivot_row = side_first + 4
    ws.cell(row=pivot_row, column=1, value="PIVOT TABLES (build manually)")
    ws.cell(row=pivot_row, column=1).font = Font(bold=True, size=12, color="C00000")
    ws.cell(row=pivot_row, column=1).fill = YELLOW_FILL
    ws.cell(row=pivot_row + 1, column=1, value=(
        "Build 3 pivots off the Trades sheet here (Insert > PivotTable): "
        "(1) volume by market, (2) volume by day x side, (3) whale count by market. "
        "Add slicers on market & side. See README manual checklist."))
    ws.cell(row=pivot_row + 1, column=1).font = NOTE_FONT

    widths = {"A": 46, "B": 16, "C": 12, "D": 12, "E": 14, "F": 10, "G": 10,
              "H": 10, "I": 10, "J": 13, "K": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return {
        "thin_range": f"E{first}:E{last_market_row}",
        "price_range": f"J{first}:J{last_market_row}",
        "day_date_range": (day_first, day_last) if days else None,
        "day_vol_range": (day_first, day_last) if days else None,
    }


# --------------------------------------------------------------------------- #
# Sheet: Dashboard
# --------------------------------------------------------------------------- #
def build_dashboard(ws, markets, trades, m_last, t_last, analysis_ref, snapshot_iso):
    n = len(markets)
    ws["A1"] = "PREDICTION MARKET OPERATIONS TRACKER"
    ws["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws["A2"] = "Live-formula operations dashboard — Polymarket real market & trade data"
    ws["A2"].font = NOTE_FONT

    snap_date = parse_date(snapshot_iso) or date.today()
    ws["A3"] = "Snapshot (UTC):"
    ws["A3"].font = KPI_LABEL_FONT
    ws["C3"] = snap_date
    ws["C3"].number_format = FMT_DATE
    ws["C3"].font = Font(bold=True, color="C00000")
    ws["A4"] = "Snapshot timestamp:"
    ws["A4"].font = KPI_LABEL_FONT
    ws["C4"] = snapshot_iso
    ws["A5"] = "Data source:"
    ws["A5"].font = KPI_LABEL_FONT
    ws["C5"] = "Polymarket Gamma API + Data API (public, no auth)"

    # ---- KPI block --------------------------------------------------------- #
    ws["A7"] = "KEY PERFORMANCE INDICATORS"
    ws["A7"].font = SECTION_FONT
    kpi_hdr = 8
    ws.cell(row=kpi_hdr, column=1, value="metric")
    ws.cell(row=kpi_hdr, column=3, value="value")
    style_header_row(ws, kpi_hdr, 1, 3)
    ws.merge_cells(start_row=kpi_hdr, start_column=1, end_row=kpi_hdr, end_column=2)

    thin = analysis_ref["thin_range"]
    price = analysis_ref["price_range"]
    kpis = [
        ("Total tracked volume (all-time, USD)", f"=SUM(Markets!$G$2:$G${m_last})", FMT_USD),
        ("Total 24h volume (USD)", f"=SUM(Markets!$F$2:$F${m_last})", FMT_USD),
        ("Largest single-market 24h volume (USD)", f"=MAX(Markets!$F$2:$F${m_last})", FMT_USD),
        ("Markets tracked", f"=COUNTA(Markets!$A$2:$A${m_last})", FMT_INT),
        ("Sports markets tracked", '=COUNTIF(Markets!$C$2:$C$%d,TRUE)' % m_last, FMT_INT),
        ("Thin-market flags", f'=COUNTIF(Analysis!{thin},"THIN MARKET")', FMT_INT),
        ("Pricing-discrepancy flags", f'=COUNTIF(Analysis!{price},"CHECK PRICING")', FMT_INT),
        ("Whale trades", f'=COUNTIF(Trades!$K$2:$K${t_last},"WHALE")', FMT_INT),
        ("Trades tracked", f"=COUNTA(Trades!$A$2:$A${t_last})", FMT_INT),
    ]
    r0 = kpi_hdr + 1
    for i, (label, formula, fmt) in enumerate(kpis):
        r = r0 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1, value=label)
        c.font = KPI_LABEL_FONT
        c.alignment = Alignment(horizontal="left")
        v = ws.cell(row=r, column=3, value=formula)
        v.number_format = fmt
        v.font = Font(bold=True, size=12, color="1F3864")
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).border = BORDER

    thin_kpi_cell = f"C{r0 + 5}"     # Thin-market flags value
    price_kpi_cell = f"C{r0 + 6}"    # Pricing-discrepancy flags value
    # Conditional formatting: red when thin markets exist, yellow for discrepancies.
    ws.conditional_formatting.add(
        thin_kpi_cell, CellIsRule(operator="greaterThan", formula=["0"],
                                  fill=RED_FILL, font=RED_FONT))
    ws.conditional_formatting.add(
        price_kpi_cell, CellIsRule(operator="greaterThan", formula=["0"],
                                   fill=YELLOW_FILL, font=YELLOW_FONT))

    # ---- helper table for the bar chart: top 10 markets by 24h volume ------ #
    # Full titles live in a hidden column (J) so XLOOKUP can match; short labels
    # in column E drive a readable chart axis. Volumes are LIVE via XLOOKUP.
    top = markets.sort_values("volume_24h", ascending=False).head(10)
    chart_hdr = 8
    ws.cell(row=chart_hdr, column=5, value="Top 10 markets (24h vol)")
    ws.cell(row=chart_hdr, column=6, value="volume_24h")
    style_header_row(ws, chart_hdr, 5, 6)
    bar_first = chart_hdr + 1
    for i, (_, m) in enumerate(top.iterrows()):
        r = bar_first + i
        full = str(m["question"])
        short = (full[:34] + "…") if len(full) > 35 else full
        ws.cell(row=r, column=5, value=short)
        ws.cell(row=r, column=10, value=full)  # hidden full title for XLOOKUP
        ws.cell(row=r, column=6,
                value=f"=_xlfn.XLOOKUP(J{r},MktTitle,MktVol24,0)").number_format = FMT_USD
    bar_last = bar_first + len(top) - 1
    ws.column_dimensions["J"].hidden = True

    # ---- helper table for the line chart: daily trade volume --------------- #
    line_hdr = bar_last + 2
    ws.cell(row=line_hdr, column=5, value="Trade date")
    ws.cell(row=line_hdr, column=6, value="daily trade volume (USD)")
    style_header_row(ws, line_hdr, 5, 6)
    line_first = line_hdr + 1
    line_last = line_first - 1
    if analysis_ref["day_date_range"]:
        a_first, a_last = analysis_ref["day_date_range"]
        for i, ar in enumerate(range(a_first, a_last + 1)):
            r = line_first + i
            # Mirror the Analysis by-day table via live references.
            ws.cell(row=r, column=5, value=f"=Analysis!A{ar}").number_format = FMT_DATE
            ws.cell(row=r, column=6, value=f"=Analysis!B{ar}").number_format = FMT_USD
            line_last = r

    # ---- charts ------------------------------------------------------------ #
    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top 10 Markets by 24h Volume"
    bar.y_axis.title = "24h volume (USD)"
    bar.height = 9
    bar.width = 20
    data = Reference(ws, min_col=6, min_row=chart_hdr, max_row=bar_last)
    cats = Reference(ws, min_col=5, min_row=bar_first, max_row=bar_last)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.legend = None
    ws.add_chart(bar, "A19")

    if line_last >= line_first:
        line = LineChart()
        line.title = "Daily Trade Volume (top-10 sports markets)"
        line.y_axis.title = "trade volume (USD)"
        line.x_axis.title = "date"
        line.height = 9
        line.width = 20
        ldata = Reference(ws, min_col=6, min_row=line_hdr, max_row=line_last)
        lcats = Reference(ws, min_col=5, min_row=line_first, max_row=line_last)
        line.add_data(ldata, titles_from_data=True)
        line.set_categories(lcats)
        line.legend = None
        ws.add_chart(line, "A37")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 22
    ws.sheet_view.showGridLines = False


# --------------------------------------------------------------------------- #
# Optional sheet: Discrepancy (sportsbook vs Polymarket cross-market mispricing)
# --------------------------------------------------------------------------- #
def _team_from_question(q):
    """Extract the subject team from a 'Will <team> win ...?' style question."""
    import re
    m = re.match(r"\s*will\s+(.+?)\s+win\b", str(q), flags=re.IGNORECASE)
    return m.group(1).strip().lower() if m else None


def build_discrepancy(ws, markets, odds):
    """Compare sportsbook implied prob vs Polymarket price; flag gaps > 5%.

    Team-name matching (sportsbook <-> Polymarket) is resolved in Python; the
    gap and mispricing flag are LIVE Excel formulas, and the Polymarket price is
    pulled live via XLOOKUP so it recalculates with the rest of the workbook.
    """
    ws["A1"] = "DISCREPANCY — sportsbook implied prob vs Polymarket (cross-market mispricing)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Flag = |Polymarket YES − sportsbook implied| > 5%. Sportsbook probs are "
                "de-vigged (odds normalized to sum to 1). Gap & flag are live formulas.")
    ws["A2"].font = NOTE_FONT

    # Build a team -> (question, yes_price) map from Polymarket sports markets.
    team_map = {}
    for _, m in markets.iterrows():
        if not bool(m.get("is_sports")):
            continue
        team = _team_from_question(m["question"])
        if team and team not in team_map:
            team_map[team] = m["question"]

    headers = ["team", "sportsbook_implied_prob", "polymarket_yes_price",
               "gap_abs", "flag"]
    hdr = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hdr, column=i, value=h)
    style_header_row(ws, hdr, 1, len(headers))

    first = hdr + 1
    r = first
    matched = 0
    for _, o in odds.iterrows():
        team = str(o["team"]).strip().lower()
        question = team_map.get(team)
        if not question:
            continue
        ws.cell(row=r, column=1, value=o["team"])
        ws.cell(row=r, column=2,
                value=float(o["sportsbook_implied_prob"])).number_format = FMT_PRICE
        ws.cell(row=r, column=7, value=question)  # hidden full title for XLOOKUP
        ws.cell(row=r, column=3,
                value=f"=_xlfn.XLOOKUP(G{r},MktTitle,MktYes,\"\")").number_format = FMT_PRICE
        ws.cell(row=r, column=4, value=f"=ABS(C{r}-B{r})").number_format = FMT_PRICE
        ws.cell(row=r, column=5, value=f'=IF(D{r}>0.05,"MISPRICED >5%","")')
        r += 1
        matched += 1
    last = r - 1

    if matched:
        ws.conditional_formatting.add(
            f"E{first}:E{last}",
            CellIsRule(operator="equal", formula=['"MISPRICED >5%"'],
                       fill=RED_FILL, font=RED_FONT))
        ws.column_dimensions["G"].hidden = True
    else:
        ws.cell(row=first, column=1,
                value="No sportsbook team matched a Polymarket market in this snapshot.")
        ws.cell(row=first, column=1).font = NOTE_FONT

    for col, w in {"A": 26, "B": 22, "C": 22, "D": 12, "E": 16}.items():
        ws.column_dimensions[col].width = w
    return matched


# --------------------------------------------------------------------------- #
# Defined names
# --------------------------------------------------------------------------- #
def add_defined_names(wb, m_last, t_last):
    names = {
        "SnapshotDate": "Dashboard!$C$3",
        "MktTitle": f"Markets!$A$2:$A${m_last}",
        "MktCat": f"Markets!$B$2:$B${m_last}",
        "MktYes": f"Markets!$D$2:$D${m_last}",
        "MktVol24": f"Markets!$F$2:$F${m_last}",
        "MktTotVol": f"Markets!$G$2:$G${m_last}",
        "WhaleThreshold": "Analysis!$B$8",
    }
    for name, ref in names.items():
        wb.defined_names[name] = DefinedName(name, attr_text=ref)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="Build the operations workbook from CSVs.")
    ap.add_argument("--markets", help="path to markets CSV")
    ap.add_argument("--trades", help="path to trades CSV")
    args = ap.parse_args()

    markets, trades, odds, snapshot_iso = resolve_snapshot(args)
    os.makedirs(OUT_DIR, exist_ok=True)

    wb = Workbook()
    ws_markets = wb.active
    ws_markets.title = "Markets"
    ws_trades = wb.create_sheet("Trades")
    ws_analysis = wb.create_sheet("Analysis")
    ws_dashboard = wb.create_sheet("Dashboard")

    m_last = build_markets(ws_markets, markets)
    t_last = build_trades(ws_trades, trades)
    add_defined_names(wb, m_last, t_last)
    analysis_ref = build_analysis(ws_analysis, markets, trades, m_last, t_last)
    build_dashboard(ws_dashboard, markets, trades, m_last, t_last,
                    analysis_ref, snapshot_iso)

    disc_n = None
    if len(odds):
        ws_disc = wb.create_sheet("Discrepancy")
        disc_n = build_discrepancy(ws_disc, markets, odds)

    wb.active = wb.sheetnames.index("Dashboard")
    wb.save(OUT_PATH)

    print(f"Workbook written: {os.path.relpath(OUT_PATH)}")
    print(f"  snapshot : {snapshot_iso}")
    print(f"  markets  : {len(markets)} rows")
    print(f"  trades   : {len(trades)} rows")
    sheets = "Markets, Trades, Analysis, Dashboard"
    if disc_n is not None:
        sheets += f", Discrepancy ({disc_n} matched teams)"
    print(f"  sheets   : {sheets}")
    print("\nOpen in Excel (365) to recalculate all live formulas and view charts.")


if __name__ == "__main__":
    main()
